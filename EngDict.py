import re
import os           # check if a file/directory exist
import xlrd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

status = 1
vocab = dict()              # input word stored as dict with its numbers of appearance
vocab_sort = list()         # sorted vocab list w.r.t. descend order
vocab_list = list()         # all vocabulary in OxfordDict
vocab_multi = dict()        # vocab in OxfordDict with multiple appearance
if os.path.isfile('Vocab_list.xlsx') is True:
    book = xlrd.open_workbook('Vocab_list.xlsx')
    sheet = book.sheet_by_name('Sheet1')
    for i in range(sheet.nrows):
        cell_value_class = sheet.cell(i,0).value    # words in the first col
        cell_value_id = int(sheet.cell(i,1).value)  # freq in the second col
        vocab[cell_value_class] = cell_value_id



# vocab_char = ['n.','v.','-v.','Abbr.','adj.','contr.','-n.','-adj.','colloq.','var.','pl.','adv.','']

# Build dictionary
fname = 'OxfordDict.txt'
fh = open(fname)
#
# for line in fh:
#     line = line.rstrip()
#     # line = line.lower()
#     if len(line) != 0:
#         words = line.lower().split()
#         vocab_list.append(words[0])
#
# multivocab_pos = 1
# for word in vocab_list:
#     if re.search('[0-9]',word):
#         vocab_multi[word] = multivocab_pos
#     multivocab_pos += 1
# print(vocab_multi)

        # print(words[0])
# 用一个separate的list把所有含数字的单词记录进去，这个list记录单词和该单词所在的位置。一旦属于这个list里的单词
# 自动提示找到多个该单词，让用户选择使用哪个词义。

# print(vocab_list)

# input
# while status == 1:
#     vocab_in = input('Please Enter a Word: ')   # input English vocab
#     if vocab_in is '1':
#         print('Process Complete!')
#         status = 0
#     elif not vocab_in in vocab_list:
#         count = 0
#         # rexpression = '^' + vocab_in + '1|^' + vocab_in + '2|^' + vocab_in + '3|^' + vocab_in + '4'
#         rexpression = '^' + vocab_in
#         print(rexpression)
#         for k, v in list(vocab_multi.items()):
#             if re.search(rexpression,k):
#                 # 有问题，比如go出现在go1和gosh1，go会算多次。
#                 # vocab[vocab_in] = vocab.get(vocab_in,0) + 1
#                 count = count + 1
#                 print(k)
#                 # print('Multiple vocab found!')
#         if count > 0:
#             vocab[vocab_in] = vocab.get(vocab_in,0) + 1
#             print('Multiple vocab found!')
#             print(vocab_in, 'appears', count, 'times.')
#         else:
#             print('Word Not Found! Please Enter Another One.')
#     elif vocab_in is '':
#         print('Please Enter a Valid Word!')
#         continue
#     else:
#         vocab[vocab_in] = vocab.get(vocab_in,0) + 1
#
# print(vocab)
while status == 1:
    vocab_in = input('Please Enter a Word: ')   # input English vocab
    if vocab_in is '1':
        print('Process Complete!')
        status = 0
    elif not vocab_in in vocab_list:
        count = 0
        rexpression = '^' + vocab_in + '1|^' \
                        + vocab_in + '2|^' \
                        + vocab_in + '3|^' \
                        + vocab_in + '4|^' \
                        + vocab_in + '\s'
        print(rexpression)
        fname = 'OxfordDict.txt'
        fh = open(fname)
        for line in fh:
            # print('working!!!')
            line_lower = line.lower().rstrip()
            # 只跑一次是什么毛病？？？？
            if re.search(rexpression,line_lower):
                count += 1
                print(line_lower)
        if count > 1:
            vocab[vocab_in] = vocab.get(vocab_in,0) + 1
            print('Multiple vocab found!')
            print(vocab_in, 'appears', count, 'times.')
        else:
            vocab[vocab_in] = vocab.get(vocab_in,0) + 1
print(vocab)


# # Sort the list
# for key, val in list(vocab.items()):
#     vocab_sort.append((val, key))
#
# vocab_sort.sort(reverse=True)
# print(vocab_sort)
#
# # res = [lis[1] for lis in vocab_sort]
# # print(str(res))
# # for lis in vocab_sort:
# #     print(lis)
# wb = Workbook()
# wbws1 = wb.active
# wbws1.title = 'Sheet1'
# count = 1
# for lis in vocab_sort:
#     _ = wbws1.cell(column=1, row=count, value = lis[1])
#     _ = wbws1.cell(column=2, row=count, value = lis[0])
#     count += 1
#
#
# wb.save('Vocab_list.xlsx')
