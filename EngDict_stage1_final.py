import re
import os
import xlrd
import itertools as it
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def DictSearching(lineinput):
    global final_output
    vocab_char1 = ['\sn\.','\sv\.','\sv\.aux\.','\sAbbr\.','\sadj\.','\scontr\.',\
               '\svar\.','\spl\.','\sadv\.']
    vocab_char2 = ['-n\.','-v\.','-v\.aux\.','-Abbr\.','-adj\.','-contr\.',\
                   '-var\.','-pl\.','-adv\.']
    vocab_char = ['n.','v.','v.aux.','Abbr.','adj.','contr.',\
               'var.','pl.','adv.','-n.','-v.','-v.aux.','-Abbr.',\
               '-adj.','-contr.', '-var.','-pl.','-adv.']
    char_property = {}
#    retrn_value = {}
    rexpression1 = str()
    rexpression2 = str()
    for char in vocab_char1:
        rexpression1 += char + '\s|'        # regex with no '-'
    for char in vocab_char2:
        rexpression2 += char + '\s|'        # regex with '-'

    seg_split3 = []
    seg_split4 = []
    final_output = []
    lineinput = lineinput.rstrip()

    related_vocab = lineinput.split('|*| ')          # get rid of those related word
    lineinput = related_vocab[0]
    wordsplit = lineinput.split()

    for char in vocab_char:
        for word in wordsplit:
            if len(char) is len(word):
                if char in word:
                    char_property[wordsplit.index(word)] = char
    if re.search(rexpression2[0:len(rexpression2)-1],lineinput):
        # print(re.findall(rexpression2[0:len(rexpression2)-1],lineinput))
        seg_split1 = re.split(rexpression2[0:len(rexpression2)-1],lineinput)  # seg after first split
    else:
        seg_split1 = re.split(rexpression1[0:len(rexpression1)-1],lineinput)
    """sort the char_property to rearange the given word's features, such as
    word, v. meaning 1, 2, 3. n. meaning 1, 2, 3..."""
    char_property_sorted = []
    for e in char_property:
        char_property_sorted.append(e)
    char_property_sorted.sort()

    for e in seg_split1:
        seg_split3.append(re.sub('\(.*\)\s1','TAPoint',e))

    for e in seg_split3:
        seg_split4.append(re.sub('\.\s[0-9]+','.\nTAPoint',e))

    for e in seg_split4:
        if e[0] is '1':
            seg_split4[seg_split4.index(e)] = 'TAPoint' + e[1:len(e)]

    for e in seg_split4:
        replace_count = it.count(1)
        final_output.append(re.sub('TAPoint',lambda x: '{}.'.format(next(replace_count)),e))

    current_index = 1
    for e in char_property_sorted:
        final_output.insert(current_index,char_property[e])
        current_index += 2
    if len(related_vocab) > 1:
        final_output.append(related_vocab[1])
    for e in final_output:
        if final_output.index(e) is 0:
            if len(e.split()) is 1:
                print('----------------------English Dict by Guanyun Liu------------------------')
                print('Current word:', e)
            elif len(e.split()) > 1:
                print('----------------------English Dict by Guanyun Liu------------------------')
                print('Current phrase:', e)
        elif len(related_vocab) > 1 and final_output.index(e) == len(final_output) - 1:
            print('**********************************************')
            print('Additional Info about', final_output[0],'\n', e)
        else:
            print('**********************************************')
            print(e)

    return



status = 1
vocab = dict()
vocab_meaning = dict()
vocab_sort = list()
vocab_list = list()
vocab_multi = dict()
vocab_old = {}

if os.path.isfile('Vocab_list.xlsx') is True:
    book = xlrd.open_workbook('Vocab_list.xlsx')
    sheet = book.sheet_by_name('Sheet1')
    for i in range(sheet.nrows):
        # i+1的目的是除去第一行，因为第一行是headline
        if i == 0:
            continue
        else:
            cell_value_class = sheet.cell(i,1).value
            cell_value_id = int(sheet.cell(i,0).value)
            cell_value_meaning1 = sheet.cell(i,2).value
            if sheet.row_len(i) > 3:
                if sheet.row_len(i) == 4:
                    cell_value_meaning2 = sheet.cell(i,3).value
                    cell_value_meaning3 = ''
                else:
                    cell_value_meaning2 = ''
                    cell_value_meaning3 = ''
                if sheet.row_len(i) == 5:
                    cell_value_meaning2 = sheet.cell(i,3).value
                    cell_value_meaning3 = sheet.cell(i,4).value
                else:
                    cell_value_meaning2 = sheet.cell(i,3).value
                    cell_value_meaning3 = ''
            else:
                cell_value_meaning2 = ''
                cell_value_meaning3 = ''
            vocab[cell_value_class] = cell_value_id
# vocab_old = vocab 不能写成这样，因为vocab_old会随着vocab的变化而变化！！！
            vocab_old[cell_value_class] = [cell_value_id,i,cell_value_meaning1,cell_value_meaning2,\
                     cell_value_meaning3]
# vocab_old is a dictionary contained previously exist words. {'word':[出现的次数，所在行数]...}


# main procedure: if status is 1, keep running. if status is 0, stop dictionary function and go the export the spreadsheet stage
# vocab_in = input('Please Enter a Word: ')
while status == 1:
    check = 0                   # check is a variable used to check whether a input is in the Oxford dictionary.
    check2 = 0
    '''if the input is in the Oxford dictionary, check is 1, if it is not in the dictionary, check is 0'''
    line_phrase = []            # record all lines with phrases
    vocab_meaning_sub = []      # a sub dict under the vocab_meaning, it records all meanings for words appeared more than one times like still1 still2...
    vocab_in = input('Please Enter a Word: ')
    if vocab_in is '1':         # if user input 1, then the shut down the dictionary function
        print('Process Complete!')
        status = 0
    elif vocab_in is '':        # if user accidentally hit the "Enter" button more than one times, it allows the user re-entering the word.
        continue
    else:
        count = 0
        # regular expression for words with number in it, such as are1, are2 ...
        rexpression = '^' + vocab_in + '1|^' \
                        + vocab_in + '2|^'   \
                        + vocab_in + '3|^'   \
                        + vocab_in + '4'
        fname = 'OxfordDict.txt'
        fh = open(fname, encoding='utf8')       # for Windows, you may need open(fname, encoding = 'utf8')
        # looping the OxfordDict.txt file to find the user input.
        # if the user input is not found, none is recorded and let the user reinput.
        for line in fh:
            line_lower = line.lower().rstrip()  # set all words in the current line to lower cases.
            if re.search(rexpression,line_lower):
                vocab[vocab_in] = vocab.get(vocab_in,0) + 1 # record the words and its' times of appearance
                count += 1                      #record the number of words, such as are1, are2 (in this case, count = 2)
                vocab_multi[count] = re.search(rexpression,line_lower).group()
                DictSearching(line)
                vocab_meaning_sub.append(final_output)
                check = 1
                check2 = 1
            # if the given input has phrases, such as still life, then this line will be stored in line_phrase list.
            elif re.search('^' + vocab_in + '\s\w+',line_lower):
                line_phrase.append(line)
                check = 1
            # this if statement is used when user input a phrase such as still life
            elif re.search('^' + vocab_in + '\s',line_lower) and len(vocab_in.split()) > 1:
                vocab[vocab_in] = vocab.get(vocab_in,0) + 1
                DictSearching(line)
                vocab_meaning_sub.append(final_output)
                check = 1
            # this if statement has to be writen at here since it can differentiate still n. and still life
            elif re.search('^' + vocab_in + '\s',line_lower):
                vocab[vocab_in] = vocab.get(vocab_in,0) + 1
                DictSearching(line)
                vocab_meaning_sub.append(final_output)
                check = 1
            else:
                continue

        if len(line_phrase) > 0:
            if len(line_phrase) is 1:
                print('There is', len(line_phrase), 'phrase founded!\n')
                user_input = input('Do you want to display it (y/n)? ')
            else:
                print('There are', len(line_phrase), 'phrases founded!\n')
                user_input = input('Do you want to display them (y/n)? ')

            if user_input is 'y':
                for element in line_phrase:
                    DictSearching(element)
        if check is 0:
            print('ERROR: Word Not Found!\nPlease Enter a Valid Word!')
        if check2 is 1:
            vocab[vocab_in] = vocab[vocab_in] - 1

        vocab_meaning[vocab_in] = vocab_meaning_sub
    # vocab_in = input('Please Enter a Word: ')
    # os.system('clear')

#        if count > 1:
##            vocab[vocab_in] = vocab.get(vocab_in,0) + 1
#            print('Multiple vocab found!')
#            print(vocab_in, 'appears', count, 'times.')

for key, val in list(vocab.items()):
    vocab_sort.append((val, key))
vocab_sort.sort(reverse=True)

wb = Workbook()
wbws1 = wb.active
wbws1.title = 'Sheet1'

row_count = 2

# Spreadsheet headline
_ = wbws1.cell(column = 1, row = 1, value = 'Search Frequency')
_ = wbws1.cell(column = 2, row = 1, value = 'Vocab')

for lis in vocab_sort:
    col_count = 3
    if lis[1] in vocab_old:
        _ = wbws1.cell(column = 2, row = row_count, value = lis[1])
#        只改出现的次数 -- 不行，xlsx file不能实现只改部分data
        _ = wbws1.cell(column = 1, row = row_count, value = lis[0])
        _ = wbws1.cell(column = 3, row = row_count, value = vocab_old[lis[1]][2])
        _ = wbws1.cell(column = 4, row = row_count, value = vocab_old[lis[1]][3])
        _ = wbws1.cell(column = 5, row = row_count, value = vocab_old[lis[1]][4])
    else:
        _ = wbws1.cell(column = 2, row = row_count, value = lis[1])
        _ = wbws1.cell(column = 1, row = row_count, value = lis[0])
        for element in vocab_meaning[lis[1]]:
            wbws1output = ''
            wbws1output_final = ''
            if len(element) > 1:
                for item in element:
                    wbws1output = wbws1output + item + '\n'
                wbws1output_final = wbws1output
            else:
                wbws1output_final = element
            _ = wbws1.cell(column = col_count, row = row_count, value = wbws1output_final)
            col_count += 1
    row_count += 1


wb.save('Vocab_list.xlsx')

wb.close()
