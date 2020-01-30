import re
import os
import xlrd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

status = 1
vocab = dict()
vocab_sort = list()
vocab_list = list()
vocab_multi = dict()

if os.path.isfile('Vocab_list.xlsx') is True:
    book = xlrd.open_workbook('Vocab_list.xlsx')
    sheet = book.sheet_by_name('Sheet1')
    for i in range(sheet.nrows):
        cell_value_calss = sheet.cell(i,0).value
        cell_value_id = int(sheet.cell(i,1).value)
        vocab[cell_value_calss] = cell_value_id
        
        
while status == 1:
    vocab_in = input('Please Enter a Word: ')
    if vocab_in is '1':
        print('Process Complete!')
        status = 0
    elif not vocab_in in vocab_list:
        count = 0
        rexpression = '^' + vocab_in + '1|^' \
                        + vocab_in + '2|^'   \
                        + vocab_in + '3|^'   \
                        + vocab_in + '4|^'   \
                        + vocab_in + '\s'
        print(rexpression)
        fname = 'OxfordDict.txt'
        fh = open(fname, encoding='utf8')       # for Windows, you may need open(fname, encoding = 'utf8')
        for line in fh:
            line_lower = line.lower().rstrip()
            if re.search(rexpression,line_lower):
                count += 1
                print(line_lower)
        if count > 1:
            vocab[vocab_in] = vocab.get(vocab_in,0) + 1
            print('Multiple vocab found!')
            print(vocab_in, 'appears', count, 'times.')
        else:
            vocab[vocab_in] = vocab.get(vocab_in,0) + 1
print(vocab)

for key, val in list(vocab.items()):
    vocab_sort.append((val, key))
vocab_sort.sort(reverse=True)
print(vocab_sort)

wb = Workbook()
wbws1 = wb.active
wbws1.title = 'Sheet1'
count = 1
for lis in vocab_sort:
    _ = wbws1.cell(column = 1, row = count, value = lis[1])
    _ = wbws1.cell(column = 2, row = count, value = lis[0])
    _ = wbws1.cell(column = 3, row = count, value = '1. abc\r\n2. abc\n')
    count += 1
    
wb.save('Vocab_list.xlsx')
    